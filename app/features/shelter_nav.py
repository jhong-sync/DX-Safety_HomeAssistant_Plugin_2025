"""
Shelter navigation feature for DX-Safety.

This module provides shelter navigation functionality
with nearest shelter calculation and mobile app notifications.
"""

import os
import csv
import urllib.parse
from typing import List, Dict, Optional, Tuple
import openpyxl
from app.common.geo import haversine_distance
from app.adapters.homeassistant.client import HAClient
from app.observability.logging_setup import get_logger

log = get_logger("dxsafety.shelter")

Shelter = Dict[str, str | float]

def load_shelters(path: str) -> List[Shelter]:
    """대피소 데이터를 파일에서 로드합니다."""
    ext = os.path.splitext(path)[1].lower()
    rows: List[Shelter] = []
    
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8") as f:
            for r in csv.DictReader(f):
                rows.append({
                    "name": r["name"], 
                    "address": r.get("address", ""),
                    "lat": float(r["lat"]), 
                    "lon": float(r["lon"])
                })
    elif ext in (".xlsx", ".xls"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        
        log.info(f"엑셀 헤더 확인: {headers}")
        
        # 정확한 컬럼명 매핑
        name_col = "Facility Name"
        lat_col = "Latitude (EPSG4326)"
        lon_col = "Longitude (EPSG4326)"
        address_col = "Lot-based Full Address"
        
        # 필수 컬럼 검증
        if name_col not in idx:
            raise ValueError(f"시설명 컬럼을 찾을 수 없습니다: {name_col}. 사용 가능한 컬럼: {list(idx.keys())}")
        
        if lat_col not in idx:
            raise ValueError(f"위도 컬럼을 찾을 수 없습니다: {lat_col}. 사용 가능한 컬럼: {list(idx.keys())}")
        
        if lon_col not in idx:
            raise ValueError(f"경도 컬럼을 찾을 수 없습니다: {lon_col}. 사용 가능한 컬럼: {list(idx.keys())}")
        
        log.info(f"컬럼 매핑 완료 - 시설명:{name_col}, 위도:{lat_col}, 경도:{lon_col}, 주소:{address_col}")
        
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 빈 행 건너뛰기
                if not row[idx[name_col]] or row[idx[name_col]] is None:
                    continue
                
                # 위도/경도 값 검증
                lat_val = row[idx[lat_col]]
                lon_val = row[idx[lon_col]]
                
                # 빈 값이나 유효하지 않은 값 건너뛰기
                if not lat_val or not lon_val or lat_val == '' or lon_val == '':
                    log.warning(f"행 {row_num} 위도/경도 값이 비어있음: {row[idx[name_col]]}")
                    continue
                
                # 숫자 변환 시도
                try:
                    lat = float(lat_val)
                    lon = float(lon_val)
                except (ValueError, TypeError):
                    log.warning(f"행 {row_num} 위도/경도 변환 실패: lat={lat_val}, lon={lon_val}")
                    continue
                
                # 유효한 좌표 범위 검증 (한국 지역)
                if not (33.0 <= lat <= 38.5 and 124.0 <= lon <= 132.0):
                    log.warning(f"행 {row_num} 좌표가 한국 지역 범위를 벗어남: lat={lat}, lon={lon}")
                    continue
                
                rows.append({
                    "name": str(row[idx[name_col]]).strip(),
                    "address": str(row[idx[address_col]]).strip() if address_col in idx and row[idx[address_col]] else "",
                    "lat": lat,
                    "lon": lon
                })
            except (ValueError, TypeError, IndexError) as e:
                log.warning(f"행 {row_num} 데이터 변환 오류 건너뜀: {row} error:{e}")
                continue
    else:
        raise ValueError("지원하지 않는 파일 형식")
    
    log.info(f"대피소 데이터 로드됨 path:{path} count:{len(rows)}")
    return rows

def build_naver_url(dlat: float, dlng: float, dname: str, appname: str) -> str:
    """네이버 지도 길찾기 URL을 생성합니다."""
    return (f"nmap://navigation?dlat={dlat:.6f}&dlng={dlng:.6f}"
            f"&dname={urllib.parse.quote(dname)}&appname={appname}")

def find_nearest(lat: float, lon: float, shelters: List[Shelter]) -> Tuple[Shelter, float]:
    """가장 가까운 대피소를 찾습니다."""
    best: Optional[Tuple[Shelter, float]] = None
    
    for s in shelters:
        d = haversine_distance(lat, lon, float(s["lat"]), float(s["lon"]))
        if best is None or d < best[1]:
            best = (s, d)
    
    if best is None:
        raise ValueError("대피소 데이터가 없습니다")
    
    return best

class ShelterNavigator:
    """대피소 네비게이션 클래스"""
    
    def __init__(self, ha: HAClient, path: str, appname: str):
        """
        초기화합니다.
        
        Args:
            ha: Home Assistant 클라이언트
            path: 대피소 데이터 파일 경로
            appname: 네이버 지도 앱 이름
        """
        self.ha = ha
        self.path = path
        self.appname = appname
        self._shelters: List[Shelter] = []
        
        log.info(f"ShelterNavigator 초기화됨 path:{path} appname:{appname}")
    
    def load(self):
        """대피소 데이터를 로드합니다."""
        self._shelters = load_shelters(self.path)
    
    async def notify_all_devices(self, notify_group: str | None = None):
        """모든 디바이스에 가까운 대피소 알림을 발송합니다."""
        if not self._shelters:
            self.load()
        
        svcs = set(await self.ha.list_notify_mobile_services())
        devices = await self.ha.get_device_trackers()
        
        log.info(f"디바이스 알림 시작 devices:{len(devices)} services:{len(svcs)}")
        
        for d in devices:
            slug = d["entity_id"].split(".", 1)[1]
            cand = f"mobile_app_{slug}"
            service = cand if cand in svcs else notify_group
            
            if not service:
                log.warning(f"알림 서비스를 찾을 수 없음 device:{d['entity_id']}")
                continue
            
            try:
                near, dist = find_nearest(d["lat"], d["lon"], self._shelters)
                url = build_naver_url(
                    float(near["lat"]), 
                    float(near["lon"]),
                    str(near["name"]), 
                    self.appname
                )
                
                title = "[대피] 가까운 대피소 안내"
                msg = f"{near['name']} ({dist:.2f}km) - 가장 가까운 대피소로 이동하세요."
                
                # 소리 설정 (긴급 알림)
                sound = {
                    "name": "Siren.wav",
                    "critical": 1,
                    "volume": 1
                }
                
                # 액션 버튼 설정
                actions = [
                    {
                        "action": "URI",
                        "title": "네이버 지도 길안내",
                        "uri": url
                    },
                    {
                        "action": "URI", 
                        "title": "구글 지도 길안내",
                        "uri": f"google.navigation:q={near['lat']},{near['lon']}&mode=w"
                    }
                ]
                
                await self.ha.notify(service, title, msg, url, sound=sound, actions=actions)
                log.info(f"대피소 알림 발송됨 device:{d['name']} shelter:{near['name']} distance:{dist:.2f}km")
                
            except Exception as e:
                log.error(f"대피소 알림 발송 실패 device:{d['entity_id']} error:{str(e)}")
                continue
